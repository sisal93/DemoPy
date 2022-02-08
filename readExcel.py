import openpyxl

path = "C:\\Users\\Saurabh\\Desktop\\saurabh.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.active

rows = sheet.max_row
cols = sheet.max_column

# cell_user = sheet.cell(row=1, column=1).value
# cell_pass = sheet.cell(row=1, column=2).value
# print(cell_user)
# print(cell_pass)

for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(row=r, column=c).value, end="      ")

    print(" ")
